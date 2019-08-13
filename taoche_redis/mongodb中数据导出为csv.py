import openpyxl
from openpyxl import load_workbook
from pymongo import MongoClient
import csv
import pandas as pd
import xlwt
"""
user 用户名
pwd 密码
server 服务器ip
port 数据库端口号
db_name 数据库名
table_name 集合名
form_data 想要获取的字段列表
"""
def export(db_name, table_name, form_data):
    # uri = 'mongodb://' + user + ':' + pwd + '@' + server + ':' + port +'/'
    client = MongoClient(host="127.0.0.1",port=27017)
    db = client[db_name][table_name]
    with open(f"{db_name}_{table_name}.csv", "w", newline='', encoding='utf-8') as csvfilewrite:
        write = csv.writer(csvfilewrite)
        write.writerow(form_data)
        allRecordRes = db.find()


        for record in allRecordRes:
            recordValueLst = []
            for field in form_data:
                if field not in record:
                    recordValueLst.append("None")
                else:
                    recordValueLst.append(record[field])
            try:
                write.writerow(recordValueLst)
            except Exception as e:
                print(f"write csv exception.e = {e}")

def csv_to_xlsx(db_name, table_name, csvfile, outfile):
    # 创建工作簿对象
    work_book = openpyxl.Workbook()
    # 创建sheet
    work_sheet = work_book.create_sheet(title=f"{db_name}_{table_name}")
    # 打开csv文件
    csvfile = open(csvfile, encoding='utf-8')
    # 获取csv.reader
    lines = csv.reader(csvfile)
    # row
    row = 1
    # 写入从csv读取的内容 如使用了以上代码 这里行数要加一
    for line in lines:
        lin = 1
        for i in line:
            work_sheet.cell(row=row, column=lin).value = i
            lin += 1
        row += 1
    # 关闭文件
    csvfile.close()
    # 保存工作表
    work_book.save(outfile)

if __name__ == '__main__':
	# #用户名
    # user = "user"
    # #密码
    # pwd = "pwd"
    # #服务器ip
    # server = "server"
    # #服务器端口号
    # port = "port"
    # #数据库名
    # db_name = "dangdang"
    #  #集合名
    # table_name = "table_name"
    #想要获取字段的列表
    form_data = [
        'id',
        'title',
        'price',
        'new_car_price',
        'down_payment',
        'monthly_payment',
        'info',
        'displacement',
        'registration_city',
        'options',
        'car_img',
        'city',
        'color'
    ]
    file = './file.csv'
    outfile = './outfile.xlsx'
    #数据库数据导出为csv格式
    export('renrenche', 'renrenche', form_data)
    #csv格式转换成excel格式
    # csv_to_xlsx('dangdang', 'dangdang', file, outfile)
